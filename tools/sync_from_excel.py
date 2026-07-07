#!/usr/bin/env python3
"""Seed/refresh the Google Sheet with what's already in Budget 2026.xlsx.

Reads every visible month's Input blocks (and income actuals) from the
workbook and adds anything the Sheet doesn't have yet, marked imported so
tools/import_to_excel.py never writes it back. Dedup is by content
(category + date + amount, counted as a multiset), so hand-typed workbook
rows, app-imported rows, and repeat runs all coexist safely. Run it
whenever you've typed entries straight into Excel; run order vs the import
script doesn't matter.

By default only the CURRENT month is synced (his call: preloaded history =
the current month's Input tab). Pass --all-months to seed everything.

Usage:
  python3 tools/sync_from_excel.py --dry-run
  python3 tools/sync_from_excel.py
  python3 tools/sync_from_excel.py --all-months
  python3 tools/sync_from_excel.py --mock-sheet f.json --out plan.json  # testing
"""
import argparse, datetime, hashlib, json, os, re, subprocess, sys

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
CONFIG_PATH = os.path.join(HERE, "config.json")
MONTH_NAMES = ["January", "February", "March", "April", "May", "June",
               "July", "August", "September", "October", "November", "December"]
EXPENSE_ROWS = range(19, 33)
D_FORMULA = re.compile(r"^=SUM\('(.+ Input)'!C(\d+):C(\d+)\)$")
EPOCH = datetime.datetime(1899, 12, 30)


def fail(msg):
    print("ABORT: " + msg)
    sys.exit(1)


def load_config():
    if not os.path.exists(CONFIG_PATH):
        fail(f"missing {CONFIG_PATH}")
    with open(CONFIG_PATH) as f:
        return json.load(f)


def curl_json(url, post_body=None):
    cmd = ["curl", "-sL", "--max-time", "120"]
    if post_body is not None:
        cmd += ["-X", "POST", "-H", "Content-Type: text/plain;charset=utf-8",
                "--data-binary", json.dumps(post_body)]
    cmd.append(url)
    out = subprocess.run(cmd, capture_output=True, text=True)
    if out.returncode != 0:
        fail("curl failed: " + out.stderr.strip()[:200])
    try:
        return json.loads(out.stdout)
    except json.JSONDecodeError:
        fail("server did not return JSON: " + out.stdout[:200])


def norm_date(v, mk):
    """Normalize a workbook Date cell to YYYY-MM-DD; fall back to the 1st."""
    if isinstance(v, datetime.datetime):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, datetime.date):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, (int, float)) and 40000 < v < 60000:  # Excel serial
        return (EPOCH + datetime.timedelta(days=float(v))).strftime("%Y-%m-%d")
    s = str(v or "").strip()
    if re.match(r"^\d{4}-\d{2}-\d{2}$", s):
        return s
    m = re.match(r"^(\d{1,2})/(\d{1,2})(?:/(\d{2,4}))?$", s)
    if m:
        mo, d = int(m.group(1)), int(m.group(2))
        y = m.group(3)
        y = int(y) + 2000 if y and len(y) == 2 else (int(y) if y else int(mk[:4]))
        try:
            return datetime.date(y, mo, d).strftime("%Y-%m-%d")
        except ValueError:
            pass
    return mk + "-01"


def month_sheets(wb):
    out = []
    for name in wb.sheetnames:
        m = re.match(r"^(" + "|".join(MONTH_NAMES) + r") (\d{4})$", name)
        if m and wb[name].sheet_state == "visible":
            mk = f"{m.group(2)}-{MONTH_NAMES.index(m.group(1)) + 1:02d}"
            out.append((name, mk))
    return out


def read_workbook(path):
    """Return (expense_rows, income_cells) from every visible month."""
    raw = openpyxl.load_workbook(path)
    data = openpyxl.load_workbook(path, data_only=True)
    expenses, incomes, warnings = [], {}, []
    for name, mk in month_sheets(raw):
        ws = raw[name]
        if ws["B33"].value != "Total":
            warnings.append(f"{name}: unexpected layout, skipped whole month")
            continue
        for r in EXPENSE_ROWS:
            cat = str(ws.cell(row=r, column=2).value)
            m = D_FORMULA.match(str(ws.cell(row=r, column=4).value))
            if not m:
                warnings.append(f"{name} row {r}: no block formula, skipped")
                continue
            iname, s, e = m.group(1), int(m.group(2)), int(m.group(3))
            for row in range(s, e + 1):
                raw_c = raw[iname].cell(row=row, column=3).value
                if raw_c is None:
                    continue
                val = data[iname].cell(row=row, column=3).value
                if val is None:
                    warnings.append(f"{iname}!C{row}: formula with no cached value, skipped "
                                    "(open + save the workbook in Excel, then re-run)")
                    continue
                try:
                    amount = round(float(val), 2)
                except (TypeError, ValueError):
                    warnings.append(f"{iname}!C{row}: non-numeric {val!r}, skipped")
                    continue
                if amount <= 0:
                    continue
                notes = data[iname].cell(row=row, column=4).value
                expenses.append({
                    "category": cat, "mk": mk,
                    "date": norm_date(raw[iname].cell(row=row, column=2).value, mk),
                    "amount": amount,
                    "notes": "" if notes is None else str(notes),
                })
        for src, r in (("Alana", 14), ("Max", 15)):
            v = data[name].cell(row=r, column=4).value
            if v is not None:
                try:
                    incomes[(mk, src)] = round(float(v), 2)
                except (TypeError, ValueError):
                    warnings.append(f"{name}!D{r}: non-numeric income {v!r}, skipped")
    return expenses, incomes, warnings


def build_additions(expenses, incomes, sheet_entries):
    """Multiset-diff workbook content against the Sheet; return entries to add."""
    def key(cat, date, amount):
        return f"{cat}|{date}|{amount:.2f}"

    sheet_count = {}
    for e in sheet_entries:
        if e["type"] != "expense":
            continue
        k = key(e["category"], str(e["date"]), round(float(e["amount"]), 2))
        sheet_count[k] = sheet_count.get(k, 0) + 1

    wb_rows = {}
    for x in expenses:
        wb_rows.setdefault(key(x["category"], x["date"], x["amount"]), []).append(x)

    additions = []
    for k, rows in sorted(wb_rows.items()):
        have = sheet_count.get(k, 0)
        for occ, x in enumerate(rows):
            if occ < have:
                continue  # already in the Sheet (app entry or previous sync)
            eid = "xl-" + hashlib.sha1(f"{k}|{occ}".encode()).hexdigest()[:16]
            additions.append({
                "id": eid, "type": "expense", "category": x["category"],
                "date": x["date"], "amount": x["amount"],
                "notes": x["notes"], "source": "", "imported": True,
            })

    income_additions = []
    for (mk, src), total in sorted(incomes.items()):
        in_sheet = sum(round(float(e["amount"]), 2) for e in sheet_entries
                       if e["type"] == "income" and str(e["date"])[:7] == mk
                       and (e.get("source") or "") == src)
        diff = round(total - in_sheet, 2)
        if diff > 0.005:
            eid = "xlinc-" + hashlib.sha1(f"{mk}|{src}|{diff:.2f}".encode()).hexdigest()[:16]
            income_additions.append({
                "id": eid, "type": "income", "category": "Income",
                "date": mk + "-01", "amount": diff,
                "notes": "from workbook", "source": src, "imported": True,
            })
        elif diff < -0.005:
            print(f"  note: {mk} {src}: Sheet income exceeds workbook by ${-diff:,.2f} "
                  "(app entries not yet imported to Excel — fine)")
    return additions, income_additions


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument("--all-months", action="store_true",
                    help="sync every visible month (default: current month only)")
    ap.add_argument("--mock-sheet", help="read Sheet entries from a JSON file (testing)")
    ap.add_argument("--out", help="write planned additions to a JSON file instead of posting")
    args = ap.parse_args()

    cfg = load_config()
    if not os.path.exists(cfg["workbook"]):
        fail("workbook not found: " + cfg["workbook"])

    if args.mock_sheet:
        with open(args.mock_sheet) as f:
            sheet_entries = json.load(f)["entries"]
    else:
        data = curl_json(f"{cfg['endpoint']}?token={cfg['token']}&action=data")
        if not data.get("ok"):
            fail("server error: " + str(data.get("error")))
        sheet_entries = data["entries"]

    expenses, incomes, warnings = read_workbook(cfg["workbook"])
    for w in warnings:
        print("  WARN " + w)
    if not args.all_months:
        cur = datetime.date.today().strftime("%Y-%m")
        expenses = [x for x in expenses if x["mk"] == cur]
        incomes = {k: v for k, v in incomes.items() if k[0] == cur}
        print(f"Scope: current month only ({cur}); use --all-months for everything")
    additions, income_additions = build_additions(expenses, incomes, sheet_entries)

    print(f"\nWorkbook rows read: {len(expenses)} expenses, {len(incomes)} income cells")
    print(f"To add to the Sheet: {len(additions)} expenses, {len(income_additions)} income entries")
    by_month = {}
    for a in additions + income_additions:
        by_month[a["date"][:7]] = by_month.get(a["date"][:7], 0) + 1
    for mk in sorted(by_month):
        print(f"  {mk}: {by_month[mk]}")

    todo = additions + income_additions
    if not todo:
        print("Sheet already has everything.")
        return
    if args.out:
        with open(args.out, "w") as f:
            json.dump(todo, f, indent=1)
        print(f"(wrote plan to {args.out}, nothing posted)")
        return
    if args.dry_run:
        print("(dry run, nothing posted)")
        return

    added = skipped = 0
    for i in range(0, len(todo), 400):
        res = curl_json(cfg["endpoint"], {"token": cfg["token"], "action": "add_many",
                                          "entries": todo[i:i + 400]})
        if not res.get("ok"):
            fail("add_many failed mid-way (safe to re-run): " + str(res.get("error")))
        added += res.get("added", 0)
        skipped += res.get("skipped", 0)
    print(f"Done: {added} added, {skipped} already present.")


if __name__ == "__main__":
    main()
