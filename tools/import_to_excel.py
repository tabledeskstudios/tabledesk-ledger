#!/usr/bin/env python3
"""Pull TableDesk Ledger entries from the Google Sheet into Budget 2026.xlsx.

Safety protocol (always): refuse if Excel has the workbook open, back up the
original, edit a temp copy, validate the temp copy, swap only if validation
passes, and only then mark the entries as imported in the Sheet.

Usage:
  python3 tools/import_to_excel.py             # interactive: shows plan, asks y/n
  python3 tools/import_to_excel.py --dry-run   # show the plan, touch nothing
  python3 tools/import_to_excel.py --yes       # no prompt
  python3 tools/import_to_excel.py --mock f.json   # read entries from file (testing)
  python3 tools/import_to_excel.py --force-income  # overwrite non-empty income cells
  python3 tools/import_to_excel.py --mark-only tools/last_import_ids.json
                                               # re-send mark_imported after a failure

Config: tools/config.json (see config.example.json).
"""
import argparse, datetime, json, os, re, shutil, subprocess, sys, zipfile
from copy import copy

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
CONFIG_PATH = os.path.join(HERE, "config.json")
IDS_PATH = os.path.join(HERE, "last_import_ids.json")

MONTH_NAMES = ["January", "February", "March", "April", "May", "June",
               "July", "August", "September", "October", "November", "December"]

EXPENSE_ROWS = range(19, 33)   # month-tab expense rows (14 categories)
TOTAL_ROW = 33
D_FORMULA = re.compile(r"^=SUM\('(.+ Input)'!C(\d+):C(\d+)\)$")
H2_FORMULA = re.compile(r"^=SUM\(C5:C(\d+)\)$")


def fail(msg):
    print("ABORT: " + msg)
    sys.exit(1)


def load_config():
    if not os.path.exists(CONFIG_PATH):
        fail(f"missing {CONFIG_PATH} — copy config.example.json and fill it in")
    with open(CONFIG_PATH) as f:
        cfg = json.load(f)
    for key in ("endpoint", "token", "workbook"):
        if not cfg.get(key):
            fail(f"config.json is missing '{key}'")
    return cfg


def curl_json(url, post_body=None):
    cmd = ["curl", "-sL", "--max-time", "60"]
    if post_body is not None:
        # no -X POST: --data-binary already POSTs, and curl must downgrade to GET
        # when it follows Apps Script's 302 redirect (as browsers do)
        cmd += ["-H", "Content-Type: text/plain;charset=utf-8",
                "--data-binary", json.dumps(post_body)]
    cmd.append(url)
    out = subprocess.run(cmd, capture_output=True, text=True)
    if out.returncode != 0:
        fail(f"curl failed: {out.stderr.strip()[:200]}")
    try:
        return json.loads(out.stdout)
    except json.JSONDecodeError:
        fail("server did not return JSON (check endpoint URL): " + out.stdout[:200])


def fetch_entries(cfg, mock_path):
    if mock_path:
        with open(mock_path) as f:
            data = json.load(f)
    else:
        data = curl_json(f"{cfg['endpoint']}?token={cfg['token']}&action=data")
    if not data.get("ok"):
        fail("server error: " + str(data.get("error")))
    return data["entries"]


def month_sheet_name(mk):
    y, m = int(mk[:4]), int(mk[5:7])
    return f"{MONTH_NAMES[m - 1]} {y}"


def check_not_open(path):
    lock = os.path.join(os.path.dirname(path), "~$" + os.path.basename(path))
    if os.path.exists(lock):
        fail("the workbook looks open in Excel (lock file present) — close it first")


def read_month_structure(wb, sheet_name):
    """Return {category: {row, input, start, end}} for one month tab, with checks."""
    ws = wb[sheet_name]
    if ws[f"B{TOTAL_ROW}"].value != "Total":
        fail(f"{sheet_name}: expected 'Total' at B{TOTAL_ROW} — workbook layout changed, "
             "update tools/import_to_excel.py before importing")
    out = {}
    for r in EXPENSE_ROWS:
        label = ws.cell(row=r, column=2).value
        m = D_FORMULA.match(str(ws.cell(row=r, column=4).value))
        if not label or not m:
            fail(f"{sheet_name} row {r}: unexpected label/formula ({label!r})")
        out[str(label)] = {"row": r, "input": m.group(1),
                           "start": int(m.group(2)), "end": int(m.group(3))}
    return out


def build_plan(wb, entries, force_income):
    """Group unimported entries into per-sheet write plans."""
    expenses, incomes, skipped = [], [], []
    for e in entries:
        if e.get("imported"):
            if e["type"] == "income":
                incomes.append(e)  # income totals always recomputed from all entries
            continue
        (expenses if e["type"] == "expense" else incomes).append(e)

    new_income_ids = [e["id"] for e in incomes if not e.get("imported")]

    plan = {}   # sheet_name -> {"structure":…, "cats": {cat: [entries]}}
    for e in expenses:
        mk = e["date"][:7]
        name = month_sheet_name(mk)
        if name not in wb.sheetnames or wb[name].sheet_state != "visible":
            skipped.append((e, f"no visible sheet '{name}'"))
            continue
        p = plan.setdefault(name, {"cats": {}})
        p["cats"].setdefault(e["category"], []).append(e)

    for name, p in plan.items():
        p["structure"] = read_month_structure(wb, name)
        for cat in p["cats"]:
            if cat not in p["structure"]:
                fail(f"category {cat!r} has no expense row on '{name}'")

    # income: recompute month/source totals from ALL income entries
    income_totals = {}  # (sheet, source) -> total
    for e in incomes:
        mk = e["date"][:7]
        name = month_sheet_name(mk)
        if name not in wb.sheetnames or wb[name].sheet_state != "visible":
            if not e.get("imported"):
                skipped.append((e, f"no visible sheet '{name}'"))
            continue
        src = (e.get("source") or "").strip()
        if src not in ("Alana", "Max"):
            if not e.get("imported"):
                skipped.append((e, f"income source {src!r} has no row (use Alana or Max)"))
            continue
        income_totals[(name, src)] = income_totals.get((name, src), 0) + float(e["amount"])

    income_writes, income_conflicts = [], []
    for (name, src), total in sorted(income_totals.items()):
        row = 14 if src == "Alana" else 15
        cur = wb[name].cell(row=row, column=4).value
        total = round(total, 2)
        try:
            cur_num = None if cur is None else round(float(cur), 2)
        except (TypeError, ValueError):
            cur_num = "non-numeric"
        if cur_num == total:
            continue
        if cur is None or (force_income and cur_num != "non-numeric"):
            income_writes.append({"sheet": name, "row": row, "src": src,
                                  "total": total, "was": cur})
        else:
            income_conflicts.append((name, src, cur, total))
    return plan, income_writes, income_conflicts, skipped, new_income_ids


def block_capacity(ws, start, end):
    empty = [r for r in range(start, end + 1) if ws.cell(row=r, column=3).value is None]
    return empty


def apply_plan(wb, plan):
    """Write expense entries into input blocks, inserting rows where blocks are full.
    Returns (writes, sheet_meta) for validation."""
    writes = []       # (input_sheet, row, date, amount, notes, month_sheet, cat)
    sheet_meta = {}   # input_sheet -> {"blocks": {cat: (final_start, final_end)}, "h2_end", "month"}

    for month_name, p in sorted(plan.items()):
        struct = p["structure"]
        input_name = next(iter(struct.values()))["input"]
        iws = wb[input_name]

        blocks = sorted(struct.items(), key=lambda kv: kv[1]["start"])
        inserts = {}  # cat -> k
        # bottom-up: insert + write, original coordinates stay valid at write time
        for cat, info in sorted(struct.items(), key=lambda kv: -kv[1]["start"]):
            entries = p["cats"].get(cat, [])
            if not entries:
                continue
            if iws.cell(row=info["start"] - 2, column=2).value != cat:
                fail(f"{input_name}: expected block header {cat!r} at "
                     f"B{info['start'] - 2} — layout mismatch")
            empty = block_capacity(iws, info["start"], info["end"])
            k = max(0, len(entries) - len(empty))
            if k:
                iws.insert_rows(info["end"] + 1, k)
                for rr in range(info["end"] + 1, info["end"] + 1 + k):
                    for col in range(2, 6):
                        dst = iws.cell(row=rr, column=col)
                        dst._style = copy(iws.cell(row=info["end"], column=col)._style)
                empty += list(range(info["end"] + 1, info["end"] + 1 + k))
                inserts[cat] = k
            entries_sorted = sorted(entries, key=lambda e: e["date"])
            for e, row in zip(entries_sorted, empty):
                iws.cell(row=row, column=2).value = e["date"]
                iws.cell(row=row, column=3).value = round(float(e["amount"]), 2)
                if e.get("notes"):
                    iws.cell(row=row, column=4).value = str(e["notes"])[:200]
                writes.append((input_name, row, e, month_name))

        # final layout: block with orig start s shifts by k of every inserted block ending above it
        final = {}
        for cat, info in blocks:
            shift = sum(k for c2, k in inserts.items()
                        if struct[c2]["end"] < info["start"])
            final[cat] = (info["start"] + shift, info["end"] + shift + inserts.get(cat, 0))
        # writes done at original coords also shift if a block above them was extended
        adjusted = []
        for (iname, row, e, mname) in writes:
            if iname != input_name:
                adjusted.append((iname, row, e, mname))
                continue
            shift = sum(k for c2, k in inserts.items()
                        if struct[c2]["end"] < row and e["category"] != c2)
            adjusted.append((iname, row + shift, e, mname))
        writes = adjusted

        mws = wb[month_name]
        for cat, (fs, fe) in final.items():
            mws.cell(row=struct[cat]["row"], column=4).value = f"=SUM('{input_name}'!C{fs}:C{fe})"
        max_end = max(fe for _, fe in final.values())
        iws["H2"] = f"=SUM(C5:C{max_end})"
        sheet_meta[input_name] = {"blocks": final, "h2_end": max_end, "month": month_name}
    return writes, sheet_meta


def validate(tmp_path, writes, sheet_meta, income_writes):
    bad = zipfile.ZipFile(tmp_path).testzip()
    if bad:
        fail(f"temp workbook corrupt: {bad}")
    wb = openpyxl.load_workbook(tmp_path)
    errs = []
    for (iname, row, e, mname) in writes:
        iws = wb[iname]
        got_amt = iws.cell(row=row, column=3).value
        if got_amt is None or abs(float(got_amt) - float(e["amount"])) > 0.005:
            errs.append(f"{iname}!C{row}: {got_amt!r} != {e['amount']}")
        if iws.cell(row=row, column=2).value != e["date"]:
            errs.append(f"{iname}!B{row}: {iws.cell(row=row, column=2).value!r} != {e['date']!r}")
    for iname, meta in sheet_meta.items():
        iws, mws = wb[iname], wb[meta["month"]]
        struct = read_month_structure(wb, meta["month"])
        for cat, (fs, fe) in meta["blocks"].items():
            if (struct[cat]["start"], struct[cat]["end"]) != (fs, fe):
                errs.append(f"{meta['month']} {cat}: formula range {struct[cat]['start']}:{struct[cat]['end']} != planned {fs}:{fe}")
            if iws.cell(row=fs - 2, column=2).value != cat:
                errs.append(f"{iname}!B{fs - 2}: block header moved (expected {cat!r})")
        if str(iws["H2"].value) != f"=SUM(C5:C{meta['h2_end']})":
            errs.append(f"{iname}!H2: {iws['H2'].value!r}")
        if str(iws["H3"].value) != f"='{meta['month']}'!D{TOTAL_ROW}":
            errs.append(f"{iname}!H3: {iws['H3'].value!r}")
        for col in "CDE":
            want = f"=SUBTOTAL(109,{col}19:{col}32)"
            if mws[f"{col}{TOTAL_ROW}"].value != want:
                errs.append(f"{meta['month']}!{col}{TOTAL_ROW}: {mws[f'{col}{TOTAL_ROW}'].value!r}")
    for w in income_writes:
        got = wb[w["sheet"]].cell(row=w["row"], column=4).value
        if got is None or abs(float(got) - w["total"]) > 0.005:
            errs.append(f"{w['sheet']}!D{w['row']}: {got!r} != {w['total']}")
    return errs


def mark_imported(cfg, ids):
    if not ids:
        return True
    res = curl_json(cfg["endpoint"], {"token": cfg["token"], "action": "mark_imported", "ids": ids})
    return bool(res.get("ok"))


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument("--yes", action="store_true")
    ap.add_argument("--mock")
    ap.add_argument("--force-income", action="store_true")
    ap.add_argument("--mark-only")
    args = ap.parse_args()

    cfg = load_config()

    if args.mark_only:
        with open(args.mark_only) as f:
            ids = json.load(f)["ids"]
        ok = mark_imported(cfg, ids)
        print("marked" if ok else "MARK FAILED — try again later")
        sys.exit(0 if ok else 1)

    wb_path = cfg["workbook"]
    if not os.path.exists(wb_path):
        fail(f"workbook not found: {wb_path}")
    check_not_open(wb_path)
    mtime_before = os.path.getmtime(wb_path)

    entries = fetch_entries(cfg, args.mock)
    wb = openpyxl.load_workbook(wb_path)
    plan, income_writes, income_conflicts, skipped, new_income_ids = build_plan(
        wb, entries, args.force_income)

    n_exp = sum(len(v) for p in plan.values() for v in p["cats"].values())
    print(f"\n=== Import plan ===")
    for name, p in sorted(plan.items()):
        for cat, es in sorted(p["cats"].items()):
            total = sum(float(e["amount"]) for e in es)
            print(f"  {name:16} {cat:26} {len(es):3} entries  ${total:,.2f}")
    for w in income_writes:
        was = "" if w["was"] is None else f" (replacing {w['was']})"
        print(f"  {w['sheet']:16} income {w['src']:19} -> ${w['total']:,.2f}{was}")
    for name, src, cur, total in income_conflicts:
        print(f"  CONFLICT {name} {src}: workbook has {cur}, app total is ${total:,.2f} "
              "(re-run with --force-income to overwrite)")
    for e, why in skipped:
        print(f"  SKIP {e['date']} {e['category']} ${float(e['amount']):,.2f}: {why}")
    if n_exp == 0 and not income_writes:
        print("Nothing to import.")
        return
    if args.dry_run:
        print("(dry run, nothing written)")
        return
    if not args.yes:
        if input("Proceed? [y/N] ").strip().lower() != "y":
            print("Cancelled.")
            return

    writes, sheet_meta = apply_plan(wb, plan)
    for w in income_writes:
        wb[w["sheet"]].cell(row=w["row"], column=4).value = w["total"]

    tmp_path = os.path.join(HERE, "_import_tmp.xlsx")
    wb.save(tmp_path)
    errs = validate(tmp_path, writes, sheet_meta, income_writes)
    if errs:
        os.remove(tmp_path)
        fail("validation failed, original untouched:\n  " + "\n  ".join(errs[:20]))

    if os.path.getmtime(wb_path) != mtime_before:
        os.remove(tmp_path)
        fail("workbook changed on disk while we worked — re-run")
    check_not_open(wb_path)

    stamp = datetime.datetime.now().strftime("%Y-%m-%d-%H%M")
    backup = os.path.join(os.path.dirname(wb_path),
                          os.path.splitext(os.path.basename(wb_path))[0] + f" BACKUP {stamp}-preimport.xlsx")
    shutil.copy2(wb_path, backup)
    shutil.move(tmp_path, wb_path)
    print(f"\nWrote {len(writes)} expense entries + {len(income_writes)} income cells.")
    print(f"Backup: {backup}")

    ids = [e["id"] for (_, _, e, _) in writes] + new_income_ids
    with open(IDS_PATH, "w") as f:
        json.dump({"ids": ids, "when": stamp}, f)
    if args.mock:
        print("(mock mode: not marking entries imported)")
        return
    if mark_imported(cfg, ids):
        print(f"Marked {len(ids)} entries as imported in the Sheet.")
    else:
        print("WARNING: could not mark entries as imported. Do NOT re-run the import;")
        print(f"first run: python3 tools/import_to_excel.py --mark-only {IDS_PATH}")


if __name__ == "__main__":
    main()
